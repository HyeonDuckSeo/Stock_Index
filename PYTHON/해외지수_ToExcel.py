import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
import pandas as pd
import xlwings as xw
import os
from datetime import datetime
import matplotlib.pyplot as plt
from datetime import datetime


###################################################################################################################################
# 예상안
################################################################################################################################### 
def expectation_toexcel(eval_dt, index_present_exp, index_expectation, exp_fin, exp_scoring):
    
    data_frame1 = index_present_exp # 현재 지수구성종목 비중
    data_frame2 = index_expectation # 예상안 
    data_frame3 = exp_fin           # 재무데이터 스크리닝
    data_frame4 = exp_scoring       # 예탁원 스코어링

    today = datetime.today().strftime("%Y%m%d") 
    excel_file_path = f'//00.000.00.00/FnGuide/퀀트본부/인덱스개발팀/06.개인별폴더/서현덕/해외지수_OFFICE/예상안_확정안/예상안/예상안_{eval_dt}.xlsx'  # Specified File Path


    # 1.SETTING UP DATA TABLE LOCATION 
    ## Create a blank column to insert between the data frames
    blank_column1 = pd.DataFrame({'': [''] * len(data_frame1)})
    blank_column2 = pd.DataFrame({'': [''] * len(data_frame3)})

    ## Combine the data frames with the blank columns
    combined_data1 = pd.concat([data_frame1, blank_column1, data_frame2], axis=1)
    combined_data2 = pd.concat([data_frame3, blank_column2, data_frame4], axis=1)


    # 2. ADDING COMBINED_DATA2 & ITS COLUMN LABELS TO THE SHEET
    ## Get the column labels from exp_fin and exp_scoring
    column_labels = list(exp_fin.columns) + [''] + list(exp_scoring.columns)


    # 3. Save or Update
    try:
        # 엑셀 파일 열기 (기존 파일을 열고 데이터프레임을 업데이트)
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            combined_data1.to_excel(writer, sheet_name="예상안", startrow=4, startcol=1, index=False)
        # print("데이터프레임을 엑셀 파일에 업데이트했습니다.")

    except FileNotFoundError:
        # 파일이 존재하지 않으면 새로운 엑셀 파일로 저장
        combined_data1.to_excel(excel_file_path, sheet_name="예상안", startrow=4, startcol=1, index=False)
        # print("새로운 엑셀 파일로 데이터프레임을 저장했습니다.")


    ## Load the workbook using openpyxl
    workbook = openpyxl.load_workbook(excel_file_path)
    worksheet = workbook.active

    ## Write the new combined data to the worksheet, starting from B21
    start_row = 21
    for r_idx, row in enumerate(combined_data2.values):
        for c_idx, value in enumerate(row):
            worksheet.cell(row=r_idx + start_row, column=c_idx + 2, value=value)
            
    ## Write the column labels to row 20
    for c_idx, label in enumerate(column_labels):
        worksheet.cell(row=20, column=c_idx + 2, value=label)
        worksheet.cell(row=20, column=c_idx + 2).alignment = Alignment(horizontal='center', vertical='center')


    # 3. FORMATTING 

    # Apply font and size
    font = Font(name='맑은 고딕', size=9)

    # Define the light grey highlight 
    light_grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Define the light red highlight for "편출"
    light_red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    # Define the pastel light green highlight for "편입"
    pastel_light_green_fill = PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid")

    # Define thin black border
    thin_black_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))


    for sheet in workbook:
        for row in sheet.iter_rows(min_row=3, max_row=3):
            for cell in row:
                cell.font = font

    # Apply font and size to the rest of the workbook            
    for sheet in workbook:
        for row in sheet.iter_rows(min_row=4):
            for cell in row:
                cell.font = font

    for col in range(2,27): 
        cell = worksheet.cell(row=20, column=col)
        cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
            
    # Remove top and bottom black borders on cell N5
    border = Border(top=Side(style=None), bottom=Side(style=None))  # Remove top and bottom borders
    workbook.active.cell(row=5, column=14).border = border

    # Remove top and bottom black borders on cell N20
    border = Border(top=Side(style=None), bottom=Side(style=None))  # Remove top and bottom borders
    workbook.active.cell(row=20, column=14).border = border

    # Apply text wrapping to cells M20 and Z20
    cell_M20 = worksheet.cell(row=20, column=13)  
    cell_M20.alignment = Alignment(wrap_text=True, horizontal='center', vertical ='top')

    cell_Z20 = worksheet.cell(row=20, column=26) 
    cell_Z20.alignment = Alignment(wrap_text=True, horizontal='center', vertical ='top')

    # Adjust row height for row 20
    worksheet.row_dimensions[20].height = 19  

    # Apply grey highlights to cells 2-14 and cells 15-25     
    for col_idx in range(2, 14): 
        workbook.active.cell(row=5, column=col_idx).fill = light_grey_fill

    for col_idx in range(15, 27): 
        workbook.active.cell(row=5, column=col_idx).fill = light_grey_fill

    for col_idx in range(2, 14): 
        workbook.active.cell(row=20, column=col_idx).fill = light_grey_fill

    for col_idx in range(15, 27): 
        workbook.active.cell(row=20, column=col_idx).fill = light_grey_fill

    # Apply light red highlight to cells with "편출"
    for row in workbook.active.iter_rows(min_row=4):
        for cell in row:
            if cell.value == "편출":
                cell.fill = light_red_fill
                cell.alignment = Alignment(horizontal='center')

    # Apply light green highlight to cells with "편입"
    for row in workbook.active.iter_rows(min_row=4):
        for cell in row:
            if cell.value == "편입":
                cell.fill = pastel_light_green_fill
                cell.alignment = Alignment(horizontal='center')
                
    # Apply light red highlight to cells in column "최종유니버스_부적격여부부" with value "Y"
    for row in workbook.active.iter_rows(min_row=4):
        for cell in row:
            if cell.column == 13 and cell.value == "Y":
                cell.fill = light_red_fill
                cell.alignment = Alignment(horizontal='center')

    # Apply light green highlight to cells in column "최종편입여부" with value "Y"
    for row in workbook.active.iter_rows(min_row=4):
        for cell in row:
            if cell.column == 26 and cell.value == "Y":
                cell.fill = pastel_light_green_fill
                cell.alignment = Alignment(horizontal='center')


    # 4. ADDING TITLES 

    # Convert eval_dt to datetime objects
    eval_date = datetime.strptime(eval_dt, "%Y%m%d")

    # Merge and center title "현재기준" in cells B2 to M2
    title_cell = workbook.active.cell(row=2, column=2)
    title_cell.value = "현재기준"
    title_cell.font = Font(name='맑은 고딕', size=9, bold=True)  # Make title bold
    title_cell.alignment = Alignment(horizontal='center')
    workbook.active.merge_cells(start_row=2, start_column=2, end_row=2, end_column=13)

    # Merge and center title "정기변경예상안" in cells O2 to Z2 w/ eval_dt
    eval_date_formatted = eval_date.strftime("%Y년 %m월")
    title_cell = workbook.active.cell(row=2, column=15)
    title_cell.value = f"{eval_date_formatted} 정기변경예상안"
    title_cell.font = Font(name='맑은 고딕', size=9, bold=True)  # Make title bold
    title_cell.alignment = Alignment(horizontal='center')
    workbook.active.merge_cells(start_row=2, start_column=15, end_row=2, end_column=26)

    # Merge and center title "종가기준" with year, month, and day (iif_dt) in cell B4, centered to the right
    title_text = f"{eval_date.strftime('%Y.%m.%d')} 종가기준"
    title_cell = workbook.active.cell(row=4, column=2)
    title_cell.value = title_text
    title_cell.font = Font(name='맑은 고딕', size=9, bold=True)  # Make title bold
    title_cell.alignment = Alignment(horizontal='right')
    workbook.active.merge_cells(start_row=4, start_column=2, end_row=4, end_column=13)

    # Merge and center title "종가기준" with year and month (eval_dt) in cell O4, centered to the right
    title_text = f"{eval_date.strftime('%Y.%m.%d')} 종가기준"
    title_cell = workbook.active.cell(row=4, column=15)
    title_cell.value = title_text
    title_cell.font = Font(name='맑은 고딕', size=9, bold=True)  # Make title bold
    title_cell.alignment = Alignment(horizontal='right')
    workbook.active.merge_cells(start_row=4, start_column=15, end_row=4, end_column=26)

    # ------- COMBINED_DATA2 TITLES 

    # Merge and center title "종목선정근거(1)_재무스크리닝" in cells B17 to M17
    title_cell = workbook.active.cell(row=17, column=2)
    title_cell.value = "종목선정근거(1)_재무스크리닝"
    title_cell.font = Font(name='맑은 고딕', size=9, bold=True)  # Make title bold
    title_cell.alignment = Alignment(horizontal='center')
    workbook.active.merge_cells(start_row=17, start_column=2, end_row=17, end_column=13)

    # Merge and center title "종목선정근거(2)_예탁원데이터스코어링" in cells O17 to Z17 w/ eval_dt
    title_cell = workbook.active.cell(row=17, column=15)
    title_cell.value = "종목선정근거(2)_예탁원데이터스코어링"
    title_cell.font = Font(name='맑은 고딕', size=9, bold=True)  # Make title bold
    title_cell.alignment = Alignment(horizontal='center')
    workbook.active.merge_cells(start_row=17, start_column=15, end_row=17, end_column=26)

    # Merge and center title "종가기준" with year, month, and day (iif_dt) in cell B19, centered to the right
    title_text = f"{eval_date.strftime('%Y.%m.%d')} 종가기준"
    title_cell = workbook.active.cell(row=19, column=2)
    title_cell.value = title_text
    title_cell.font = Font(name='맑은 고딕', size=9, bold=True)  # Make title bold
    title_cell.alignment = Alignment(horizontal='right')
    workbook.active.merge_cells(start_row=19, start_column=2, end_row=19, end_column=13)

    # Merge and center title "종가기준" with year and month (eval_dt) in cell O19, centered to the right
    title_text = f"{eval_date.strftime('%Y.%m.%d')} 종가기준"
    title_cell = workbook.active.cell(row=19, column=15)
    title_cell.value = title_text
    title_cell.font = Font(name='맑은 고딕', size=9, bold=True)  # Make title bold
    title_cell.alignment = Alignment(horizontal='right')
    workbook.active.merge_cells(start_row=19, start_column=15, end_row=19, end_column=26)
            
    ################################# SAVING FINAL RESULT #############################################

    # Save the modified workbook to the specified file path
    workbook.save(excel_file_path)
    
    return 


###################################################################################################################################
# 확정안
################################################################################################################################### 
def confirmation_toexcel(iif_dt, index_present_conf, index_confirmation):
    
    data_frame1 = index_present_conf # 현재 지수구성종목 비중
    data_frame2 = index_confirmation # 예상안 

    today = datetime.today().strftime("%Y%m%d") 
    excel_file_path = f'//00.000.00.00/FnGuide/퀀트본부/인덱스개발팀/06.개인별폴더/서현덕/해외지수_OFFICE/예상안_확정안/확정안/확정안_{iif_dt}.xlsx'  # Specified File Path


    # 1. SETTING UP DATA TABLE LOCATION 
    ## Create a blank column to insert between the data frames
    blank_column = pd.DataFrame({'': [''] * len(data_frame1)})

    ## Combine the data frames with the blank column
    combined_data = pd.concat([data_frame1, blank_column, data_frame2], axis=1)


    # 2. Save or Update
    try:
        ## 엑셀 파일 열기 (기존 파일을 열고 데이터프레임을 업데이트)
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            combined_data.to_excel(writer, sheet_name="확정안", startrow=4, startcol=1, index=False)

    except FileNotFoundError:
        ## 파일이 존재하지 않으면 새로운 엑셀 파일로 저장
        combined_data.to_excel(excel_file_path, sheet_name="확정안", startrow=4, startcol=1, index=False)

    # Load the workbook using openpyxl
    workbook = openpyxl.load_workbook(excel_file_path)
    worksheet = workbook.active

        
    # 3. FORMATTING 

    # Apply font and size
    font = Font(name='맑은 고딕', size=9)

    # Define the light grey highlight 
    light_grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Define the light red highlight for "편출"
    light_red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    # Define the pastel light green highlight for "편입"
    pastel_light_green_fill = PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid")

    # Define thin black border
    thin_black_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

    for sheet in workbook:
        for row in sheet.iter_rows(min_row=3, max_row=3):
            for cell in row:
                cell.font = font

    # Apply font and size to the rest of the workbook            
    for sheet in workbook:
        for row in sheet.iter_rows(min_row=4):
            for cell in row:
                cell.font = font

    # for col in range(2,27): 
    #     cell = worksheet.cell(row=20, column=col)
    #     cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
            
    # Remove top and bottom black borders on cell N5
    border = Border(top=Side(style=None), bottom=Side(style=None))  # Remove top and bottom borders
    workbook.active.cell(row=5, column=14).border = border


    # Apply grey highlights to cells 2-14 and cells 15-25     
    for col_idx in range(2, 14): 
        workbook.active.cell(row=5, column=col_idx).fill = light_grey_fill

    for col_idx in range(15, 27): 
        workbook.active.cell(row=5, column=col_idx).fill = light_grey_fill


    # Apply light red highlight to cells with "편출"
    for row in workbook.active.iter_rows(min_row=4):
        for cell in row:
            if cell.value == "편출":
                cell.fill = light_red_fill
                cell.alignment = Alignment(horizontal='center')

    # Apply light green highlight to cells with "편입"
    for row in workbook.active.iter_rows(min_row=4):
        for cell in row:
            if cell.value == "편입":
                cell.fill = pastel_light_green_fill
                cell.alignment = Alignment(horizontal='center')
                
    # Apply light red highlight to cells in column "최종유니버스_부적격여부부" with value "Y"
    for row in workbook.active.iter_rows(min_row=4):
        for cell in row:
            if cell.column == 13 and cell.value == "Y":
                cell.fill = light_red_fill
                cell.alignment = Alignment(horizontal='center')

    # Apply light green highlight to cells in column "최종편입여부" with value "Y"
    for row in workbook.active.iter_rows(min_row=4):
        for cell in row:
            if cell.column == 26 and cell.value == "Y":
                cell.fill = pastel_light_green_fill
                cell.alignment = Alignment(horizontal='center')

    ###################################### ADDING TITLES #################################################

    # Convert eval_dt to datetime objects
    iif_date = datetime.strptime(iif_dt, "%Y%m%d")

    # Merge and center title "현재기준" in cells B2 to M2
    title_cell = workbook.active.cell(row=2, column=2)
    title_cell.value = "현재기준"
    title_cell.font = Font(name='맑은 고딕', size=9, bold=True)  # Make title bold
    title_cell.alignment = Alignment(horizontal='center')
    workbook.active.merge_cells(start_row=2, start_column=2, end_row=2, end_column=13)

    # Merge and center title "정기변경예상안" in cells O2 to Z2 w/ eval_dt
    iif_date_formatted = iif_date.strftime("%Y년 %m월")
    title_cell = workbook.active.cell(row=2, column=15)
    title_cell.value = f"{iif_date_formatted} 정기변경확정안"
    title_cell.font = Font(name='맑은 고딕', size=9, bold=True)  # Make title bold
    title_cell.alignment = Alignment(horizontal='center')
    workbook.active.merge_cells(start_row=2, start_column=15, end_row=2, end_column=26)

    # Merge and center title "종가기준" with year, month, and day (iif_dt) in cell B4, centered to the right
    title_text = f"{iif_date.strftime('%Y.%m.%d')} 종가기준"
    title_cell = workbook.active.cell(row=4, column=2)
    title_cell.value = title_text
    title_cell.font = Font(name='맑은 고딕', size=9, bold=True)  # Make title bold
    title_cell.alignment = Alignment(horizontal='right')
    workbook.active.merge_cells(start_row=4, start_column=2, end_row=4, end_column=13)

    # Merge and center title "종가기준" with year and month (eval_dt) in cell O4, centered to the right
    title_text = f"{iif_date.strftime('%Y.%m.%d')} 종가기준"
    title_cell = workbook.active.cell(row=4, column=15)
    title_cell.value = title_text
    title_cell.font = Font(name='맑은 고딕', size=9, bold=True)  # Make title bold
    title_cell.alignment = Alignment(horizontal='right')
    workbook.active.merge_cells(start_row=4, start_column=15, end_row=4, end_column=26)

            
    ################################# SAVING FINAL RESULT #############################################

    # Save the modified workbook to the specified file path
    workbook.save(excel_file_path)

    return


###################################################################################################################################
# 시뮬레이션 리포트
###################################################################################################################################

def simulationreport_toexcel(financial_data, raw_universe_df, universe_df, indices, df_rtn):
    # DataFrame List
    df1 = pd.DataFrame(df_rtn)
    df2 = pd.DataFrame(indices)
    df3 = pd.DataFrame(universe_df)
    df4 = pd.DataFrame(raw_universe_df)
    df5 = pd.DataFrame(financial_data)

    # Define sheet names for the DataFrames
    dfs_with_sheet_names = {'PERFORMANCE_DATA': df1, 'INDEX_DATA': df2, 'CONSTITUENT_DATA': df3, 'SCORING_DATA':df4, 'FINANCIAL_DATA':df5}

    # Define the new sheet name mapping (if you want to rename sheets)
    sheet_name_mapping = {'PERFORMANCE DATA': df1, 'INDEX DATA': df2, 'CONSTITUENT DATA': df3, 'SCORING DATA':df4, 'FINANCIAL DATA':df5}

    # Save the workbook to the specified file path
    today = datetime.today().strftime("%Y%m%d") 
    excel_file_path = f'//00.000.00.00/FnGuide/퀀트본부/인덱스개발팀/06.개인별폴더/서현덕/해외지수_OFFICE/시뮬레이션리포트/simulationreport_{today}.xlsx'

    # 파일이 이미 존재하는 경우 삭제
    if os.path.exists(excel_file_path):
        os.remove(excel_file_path)

    wb = xw.Book()

    # Loop through each sheet in the workbook
    for sheet_name, df in dfs_with_sheet_names.items():
        # print(sheet_name)
        wb.sheets.add(name=sheet_name)
        sheet = wb.sheets[sheet_name]
        sheet.range('1:1').color = (192, 192, 192)  # Apply gray highlight to top row
        sheet.range('A1').value = df
        
        if sheet_name == 'INDEX_DATA':
            chart = sheet.charts.add()
            chart.set_source_data(sheet.range('A1').expand())
            chart.api[1].HasTitle = True
            chart.api[1].ChartTitle.Text = 'Index TimeSeries'
            chart.height = 500
            chart.width = 900
            
            chart.chart_type = 'line'
            chart.top = sheet.range('E2').top
            chart.left = sheet.range('E2').left
        
        if sheet_name == 'PERFORMANCE_DATA':
            chart = sheet.charts.add()
            chart.set_source_data(sheet.range('A1').expand())
            chart.api[1].HasTitle = True
            chart.api[1].ChartTitle.Text = 'Performance Analysis'
            chart.height = 500
            chart.width = 800

            chart.top = sheet.range('E2').top
            chart.left = sheet.range('E2').left

            # Set x-axis labels to show all TRDT labels
            chart.api[1].Axes(1, 1).TickLabels.NumberFormat = 'yyyy-mm-dd'  # Format the date
            chart.api[1].Axes(1, 1).TickLabels.Orientation = 45  # Rotate labels
            chart.api[1].Axes(1, 1).TickLabels.ReadingOrder = -5002  # Rotate labels from top to bottom
            chart.api[1].Axes(1, 1).TickLabels.Offset = 100

    # Loop through each sheet in the workbook and set font and size
    for sheet in wb.sheets:
        sheet.api.Cells.Font.Name = '맑은 고딕'
        sheet.api.Cells.Font.Size = 9

    wb.save(excel_file_path)
    wb.close()
    
    return